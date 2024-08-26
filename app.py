import sys
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

wb_movies = load_workbook(filename='moviesnames.xlsx')
ws_movies = wb_movies.active

# Iterate through all cells in column A (Movie Names) and B (Release Dates)
for movie_cell, release_date_cell in zip(ws_movies['A'], ws_movies['B']):
    movie_name = movie_cell.value
    release_date = release_date_cell.value

    if movie_name and release_date:  # Check if cells are not empty
        driver = webdriver.Chrome()
        driver.implicitly_wait(10)

        driver.get('https://www.plex.tv/')

        search_box = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'universal-search-input')))
        search_box.click()
        search_box.send_keys(movie_name)

        found = False

        for i in range(10):
            try:
                search_result_title = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, f'#universal-search-item-{i} .nav-search-result-title')))

                release_year_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, f'#universal-search-item-{i} .chroma_Text_module_secondary')))
                release_year_text = release_year_element.text
                search_result_release_year = int(release_year_text.split('•')[-1].strip())  # Extract year and convert to integer

                if search_result_release_year == release_date:
                    found = True
                    search_result_title.click()
                    break
                else:
                    print(f"Release year for '{movie_name}' in search result {i+1} is not {release_date}. Trying next search result...")

            except Exception as e:
                print(f"An error occurred while processing search result {i+1} for '{movie_name}': {e}")

        if not found:
            print(f"No search result found for '{movie_name}'")
            driver.quit()
            continue

        current_url = driver.current_url
        print("Movie URL:", current_url)

        driver.implicitly_wait(2)
        driver.save_screenshot("last_page_screenshot.png")

        print("Movie Name:", movie_name)

        add_to_watchlist_button = driver.find_elements(By.XPATH, '//span[contains(@title, "Add to Watchlist")]')
        if add_to_watchlist_button:
            movie_availability = "Movie not available"
        else:
            movie_availability = "Movie available for Play"

        today_date = datetime.today().strftime("%m/%d/%Y")
        print("Today's date:", today_date)

        if os.path.exists('movieslist.xlsx'):
            wb_list = load_workbook(filename='movieslist.xlsx')
        else:
            wb_list = Workbook()
            ws_list = wb_list.active
            ws_list.append(['Movie Name', 'Release Date', 'Availability', 'Date', 'URL', 'Image'])
            wb_list.save(filename='movieslist.xlsx')
            wb_list = load_workbook(filename='movieslist.xlsx')

        ws_list = wb_list.active

        # Append movie_name and release_date
        ws_list.append([movie_name, release_date, movie_availability, today_date, current_url])

        img = Image('last_page_screenshot.png')
        img.width = 200
        img.height = 150
        ws_list.add_image(img, 'F%d' % (ws_list.max_row))

        wb_list.save(filename='movieslist.xlsx')

driver.quit()
